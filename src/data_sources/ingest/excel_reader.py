"""Read xlsx sheets into raw text representations."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from data_sources.models import RawSheet

logger = logging.getLogger(__name__)

__all__ = ["read_excel_sheets"]


def _format_cell(value: object) -> str:
    """Convert a cell value to a display string.

    Numbers are preserved exactly as-is — no rounding, no comma formatting,
    no decimal truncation. This ensures full numeric fidelity for downstream use.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # Preserve full precision: repr() gives the shortest representation
        # that uniquely identifies the float (no rounding/truncation)
        if value == int(value):
            return str(int(value))
        return repr(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _sheet_to_text(ws) -> str:  # noqa: ANN001 – openpyxl Worksheet
    """Convert a worksheet to column-aligned text grid."""
    rows_data: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows_data.append([_format_cell(c) for c in row])

    if not rows_data:
        return ""

    # Calculate column widths for alignment
    num_cols = max(len(r) for r in rows_data)
    col_widths = [0] * num_cols
    for row in rows_data:
        for i, cell in enumerate(row):
            if i < num_cols:
                col_widths[i] = max(col_widths[i], len(cell))

    # No cap on column widths — preserve full content without truncation

    lines: list[str] = []
    for row_idx, row in enumerate(rows_data, start=1):
        parts: list[str] = []
        for i in range(num_cols):
            cell = row[i] if i < len(row) else ""
            # Left-align text, right-align numbers
            if cell and cell.replace(",", "").replace(".", "").replace("-", "").isdigit():
                parts.append(cell.rjust(col_widths[i]))
            else:
                parts.append(cell.ljust(col_widths[i]))
        line = "  ".join(parts).rstrip()
        if line:  # Skip entirely blank rows
            lines.append(f"Row {row_idx}: {line}")

    return "\n".join(lines)


def read_excel_sheets(file_path: str | Path) -> list[RawSheet]:
    """Read all sheets from an xlsx file and return raw text representations.

    Uses data_only=True to read cached formula values rather than formulas.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheets: list[RawSheet] = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        raw_content = _sheet_to_text(ws)
        sheets.append(
            RawSheet(
                sheet_index=idx,
                sheet_name=sheet_name,
                raw_content=raw_content,
            )
        )
        logger.debug(
            "Read sheet %s (%d chars)",
            sheet_name,
            len(raw_content),
            extra={"event": "sheet_read", "sheet_name": sheet_name},
        )

    wb.close()
    logger.info(
        "Read %d sheets from %s",
        len(sheets),
        path.name,
        extra={"event": "excel_read_complete", "file": path.name, "sheet_count": len(sheets)},
    )
    return sheets
